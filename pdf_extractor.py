import pandas as pd
from PyPDF2 import PdfReader
import pdfplumber
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import streamlit as st
import base64
import io
import os
import re

def extract_tables_from_pdf(pdf_file, output_excel_path=None):
    """
    Extract tables from a PDF file and optionally save them to an Excel file.
    
    Args:
        pdf_file: File object or file path
        output_excel_path (str, optional): Path to save the Excel file
    
    Returns:
        tuple: (List of DataFrames, success status)
    """
    print(f"Processing PDF...")
    
    # Create a workbook if output path is provided
    if output_excel_path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Tables"
    
    extracted_dfs = []
    
    try:
        # Using pdfplumber to extract tables
        with pdfplumber.open(pdf_file) as pdf:
            all_tables = []
            
            for page_num, page in enumerate(pdf.pages):
                print(f"Processing page {page_num + 1}/{len(pdf.pages)}...")
                
                # First, try to extract tables using built-in table detection
                tables = page.extract_tables()
                
                if tables:
                    print(f"Found {len(tables)} tables on page {page_num + 1} using built-in detection")
                    all_tables.extend(tables)
                else:
                    # If no tables found, try our custom extraction method for different formats
                    print(f"No tables found on page {page_num + 1} using built-in detection, trying custom extraction...")
                    
                    # Try bank statement extraction
                    custom_table = extract_bank_statement_table(page)
                    if custom_table:
                        all_tables.append(custom_table)
                        print(f"Found table using bank statement detection on page {page_num + 1}")
                    else:
                        # Try general text table extraction as fallback
                        custom_table = extract_text_table(page)
                        if custom_table:
                            all_tables.append(custom_table)
                            print(f"Found table using text pattern detection on page {page_num + 1}")
            
            # If we found any tables
            if all_tables:
                sheet_index = 1
                
                for table_num, table in enumerate(all_tables):
                    # Create a clean DataFrame
                    if table:
                        # Convert to DataFrame
                        df = pd.DataFrame(table)
                        
                        # Clean up the table
                        df = clean_table(df)
                        extracted_dfs.append(df)
                        
                        # If output path is provided, write to Excel
                        if output_excel_path:
                            # If this is not the first table, create a new worksheet
                            if table_num > 0:
                                ws = wb.create_sheet(f"Table {sheet_index}")
                                sheet_index += 1
                            
                            # Write DataFrame to worksheet
                            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                                for c_idx, value in enumerate(row, 1):
                                    ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Save the workbook if output path is provided
                if output_excel_path:
                    wb.save(output_excel_path)
                    print(f"Tables saved to {output_excel_path}")
                
                return extracted_dfs, True
            else:
                print("No tables found in the PDF")
                return [], False
    
    except Exception as e:
        print(f"Error extracting tables: {e}")
        return [], False

def extract_bank_statement_table(page):
    """
    Custom function to extract tables from bank statements
    
    Args:
        page: A pdfplumber page object
    
    Returns:
        list: List of lists representing table rows and columns
    """
    # Get text and its position
    text = page.extract_text()
    
    if not text:
        return None
    
    # For bank statements, we're looking for transaction lines
    lines = text.split('\n')
    
    # Regular expression to match dates (DD-MMM-YYYY format)
    date_pattern = re.compile(r'\d{2}-[A-Za-z]{3}-\d{4}')
    
    # Regular expression to match various transaction patterns
    transaction_patterns = [
        # Pattern for the sample bank statement
        re.compile(r'(\d{2}-[A-Za-z]{3}-\d{4})\s+(.*?)(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)\s+(\d{1,3}(?:,\d{3})*(?:\.\d{2})?(?:Dr|Cr)?)'),
        # Alternative pattern without the balance
        re.compile(r'(\d{2}-[A-Za-z]{3}-\d{4})\s+(.*?)(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)'),
        # Pattern for date and description only
        re.compile(r'(\d{2}-[A-Za-z]{3}-\d{4})\s+([A-Za-z].*)')
    ]
    
    # Extract account information
    account_info = extract_account_info(lines)
    
    table_rows = []
    header = ["Date", "Description", "Debit", "Credit", "Balance"]
    table_rows.append(header)
    
    # Process each line in the text
    for line in lines:
        # Skip header and footer lines
        if "BANK NAME" in line or "Page No" in line or "REPORT PRINTED BY" in line or "----------------" in line:
            continue
        
        # Try each transaction pattern
        for pattern in transaction_patterns:
            match = pattern.search(line)
            if match and date_pattern.search(line):
                date = match.group(1)
                description = match.group(2).strip()
                
                # Handle debit or credit amount
                if len(match.groups()) >= 3:
                    amount = match.group(3).strip()
                    # Determine if debit or credit
                    if "C By" in description or "Cr-" in description or "Credit" in description:
                        debit = ""
                        credit = amount
                    else:
                        debit = amount
                        credit = ""
                else:
                    debit = ""
                    credit = ""
                
                # Handle balance if available
                balance = match.group(4) if len(match.groups()) >= 4 else ""
                
                table_rows.append([date, description, debit, credit, balance])
                break
    
    # If account info was extracted, add it as metadata rows
    if account_info:
        for key, value in account_info.items():
            table_rows.insert(1, ["", key, value, "", ""])
    
    return table_rows if len(table_rows) > 1 else None

def extract_text_table(page):
    """
    Extract tabular data from text using regular spacing
    
    Args:
        page: A pdfplumber page object
    
    Returns:
        list: List of lists representing table rows and columns
    """
    # Get text
    text = page.extract_text()
    
    if not text:
        return None
    
    lines = text.split('\n')
    
    # Try to identify table structure based on character positions
    columns = []
    data_rows = []
    
    # Look for consistent spacing that might indicate columns
    for line in lines[:10]:  # Check first few lines to determine structure
        positions = [match.start() for match in re.finditer(r'\S+', line)]
        if len(positions) > 3:  # Potential table row
            columns.append(positions)
    
    if not columns:
        return None
    
    # Analyze column positions
    avg_positions = []
    if columns:
        # Transpose the positions to get columns
        all_pos = []
        max_cols = max(len(pos) for pos in columns)
        
        for i in range(max_cols):
            col_positions = [pos[i] if i < len(pos) else None for pos in columns]
            col_positions = [p for p in col_positions if p is not None]
            if col_positions:
                avg_positions.append(sum(col_positions) // len(col_positions))
    
    # If we identified potential column positions
    if avg_positions:
        for line in lines:
            if not line.strip():
                continue
                
            # Split the line according to identified column positions
            row_data = []
            last_pos = 0
            
            for pos in avg_positions:
                if pos > len(line):
                    row_data.append("")
                    continue
                    
                cell_data = line[last_pos:pos].strip()
                row_data.append(cell_data)
                last_pos = pos
                
            # Add the last cell
            if last_pos < len(line):
                row_data.append(line[last_pos:].strip())
                
            data_rows.append(row_data)
            
        return data_rows if data_rows else None
    
    return None

def extract_account_info(lines):
    """
    Extract account information from bank statement header
    
    Args:
        lines: List of text lines from the PDF
    
    Returns:
        dict: Dictionary of account information
    """
    account_info = {}
    
    # Patterns to extract common account details
    account_patterns = {
        'Account Number': re.compile(r'Account\s+No\s*:\s*([0-9]+)'),
        'Account Name': re.compile(r'A/C\s+Name\s*:\s*(.+)'),
        'Account Holder': re.compile(r'A/C\s+Holder\s*:\s*(.+)'),
        'Open Date': re.compile(r'Open\s+Date\s*:\s*(.+)'),
        'Interest Rate': re.compile(r'Interest\s+Rate\s*:\s*(.+)'),
        'Statement Period': re.compile(r'Statement\s+of\s+account\s+for\s+the\s+period\s+of\s+(.+)')
    }
    
    for line in lines:
        for key, pattern in account_patterns.items():
            match = pattern.search(line)
            if match:
                account_info[key] = match.group(1).strip()
    
    return account_info

def clean_table(df):
    """
    Clean up the extracted table DataFrame
    
    Args:
        df (pandas.DataFrame): The DataFrame to clean
    
    Returns:
        pandas.DataFrame: Cleaned DataFrame
    """
    # Replace None and empty strings with NaN
    df = df.replace(['', None], np.nan).infer_objects(copy=False)

    # Drop rows where all elements are NaN
    df = df.dropna(how='all')

    # Fill NaN with empty string
    df = df.fillna('')

    # Set column names from first row if they look like headers
    if df.shape[0] > 0:
        if all(isinstance(x, str) and x == x.upper() for x in df.iloc[0].dropna()):
            df.columns = df.iloc[0]
            df = df.drop(0)

    # Reset index
    df = df.reset_index(drop=True)

    return df

def get_table_download_link(df, filename, index=False):
    """
    Generate a link to download the DataFrame as an Excel file
    
    Args:
        df (pandas.DataFrame): DataFrame to download
        filename (str): Name of the file
        index (bool): Whether to include index in the Excel file
    
    Returns:
        str: HTML link to download the file
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=index, sheet_name='Table')
    
    b64 = base64.b64encode(output.getvalue()).decode()
    return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">Download Excel file</a>'

# Streamlit UI definition
def main():
    st.set_page_config(page_title="PDF Table Extractor", page_icon="📊", layout="wide")
    
    st.title("PDF Table Extractor")
    st.write("Upload a PDF file to extract tables without using Tabula or Camelot")
    
    # Sidebar options
    st.sidebar.header("Options")
    extraction_method = st.sidebar.radio(
        "Extraction Method",
        ["Automatic (Try all methods)", "Built-in Detection", "Bank Statement", "Text Table"]
    )
    
    # File uploader
    uploaded_file = st.file_uploader("Choose a PDF file", type="pdf")
    
    if uploaded_file is not None:
        # Save the file temporarily to process with pdfplumber
        with open("temp.pdf", "wb") as f:
            f.write(uploaded_file.getbuffer())
            
        # Process the PDF upon uploading
        if st.button("Extract Tables"):
            with st.spinner("Extracting tables..."):
                dfs, success = extract_tables_from_pdf("temp.pdf")
                
                if success:
                    st.success(f"Extracted {len(dfs)} tables from the PDF")
                    
                    # Display tabs for each extracted table
                    tabs = st.tabs([f"Table {i+1}" for i in range(len(dfs))])
                    
                    for i, (tab, df) in enumerate(zip(tabs, dfs)):
                        with tab:
                            st.dataframe(df)
                            st.markdown(get_table_download_link(df, f"table_{i+1}.xlsx"), unsafe_allow_html=True)
                    
                    # Option to download all tables as one Excel file
                    if len(dfs) > 1:
                        all_tables_excel = io.BytesIO()
                        with pd.ExcelWriter(all_tables_excel, engine='openpyxl') as writer:
                            for i, df in enumerate(dfs):
                                df.to_excel(writer, sheet_name=f"Table {i+1}", index=False)
                        
                        b64 = base64.b64encode(all_tables_excel.getvalue()).decode()
                        st.markdown(
                            f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="all_tables.xlsx">Download All Tables</a>',
                            unsafe_allow_html=True
                        )
                else:
                    st.error("No tables were found in the PDF")
            
            # Clean up temp file
            if os.path.exists("temp.pdf"):
                os.remove("temp.pdf")

if __name__ == "__main__":
    main()